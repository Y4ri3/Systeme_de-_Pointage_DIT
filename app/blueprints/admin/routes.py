import os
import secrets
import string
from pathlib import Path
from datetime import date, datetime, timedelta

from flask import current_app, jsonify, request
from flask_jwt_extended import get_jwt_identity
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename

from app.blueprints.admin import admin_bp
from app import db
from app.models.cours import Cours
from app.models.matiere import Matiere
from app.models.notification import Notification
from app.models.parametre import Parametre
from app.models.pointage import Pointage
from app.models.promotion import Promotion
from app.models.salle import Salle
from app.models.suivi_absences import SuiviAbsences
from app.models.utilisateur import Utilisateur
from app.models.filiere import Filiere
from app.services.email_service import send_temporary_password_email
from app.services.export_service import build_csv_response, build_excel_response
from app.utils import utcnow
from app.utils.decorators import role_required
from app.utils.helpers import (
    get_or_404,
    get_pagination_params,
    paginate_query,
    serialize_cours,
    serialize_filiere,
    serialize_matiere,
    serialize_notification,
    serialize_parametre,
    serialize_pointage,
    serialize_promotion,
    serialize_salle,
    serialize_suivi_absence,
    serialize_utilisateur,
)


REPORT_TEMPLATES = [
    {
        'id': 'absences_summary',
        'name': 'Suivi des absences',
        'description': "Export du suivi des absences par etudiant et par matiere, filtrable par promotion.",
        'formats': ['csv', 'xlsx'],
        'parameters': [
            {'name': 'promotion_id', 'type': 'integer', 'required': False},
        ],
    },
    {
        'id': 'course_attendance',
        'name': "Feuille de presence d'un cours",
        'description': "Export detaille de la feuille de presence d'un cours donne.",
        'formats': ['csv', 'xlsx'],
        'parameters': [
            {'name': 'cours_id', 'type': 'integer', 'required': True},
        ],
    },
]

_ABSENCES_EXPORT_HEADERS = [
    'etudiant_id', 'nom', 'prenom', 'email', 'niveau', 'filiere',
    'matiere_code', 'matiere_nom', 'nombre_absences',
    'nb_absences_justifiees', 'seuil_atteint', 'updated_at',
]

_COURSE_ATTENDANCE_EXPORT_HEADERS = [
    'etudiant_id', 'nom', 'prenom', 'email', 'statut_presence',
    'timestamp_pointage', 'methode', 'justificatif',
]


def _course_attendance_rows(cours):
    now = utcnow()
    rows = []
    summary = {
        'present': 0,
        'retard': 0,
        'invalide': 0,
        'absent': 0,
        'non_pointer': 0,
        'absence_justifiee': 0,
    }

    for etudiant in cours.promotion.etudiants:
        pointage = Pointage.query.filter_by(
            cours_id=cours.id,
            etudiant_id=etudiant.id,
        ).order_by(Pointage.timestamp_pointage.desc()).first()

        if pointage is not None:
            statut = pointage.statut
        elif datetime.combine(cours.date, cours.heure_fin) < now:
            statut = 'absent'
        else:
            statut = 'non_pointer'

        summary[statut] = summary.get(statut, 0) + 1
        rows.append({
            'id': etudiant.id,
            'nom': etudiant.nom,
            'prenom': etudiant.prenom,
            'email': etudiant.email,
            'statut_presence': statut,
            'pointage': {
                'id': pointage.id,
                'timestamp_pointage': pointage.timestamp_pointage.isoformat(),
                'statut': pointage.statut,
                'methode': pointage.methode,
                'justificatif': pointage.justificatif,
            } if pointage else None,
        })

    return rows, summary


def _absences_summary_export_rows(promotion_id=None):
    query = SuiviAbsences.query.join(SuiviAbsences.etudiant).join(SuiviAbsences.matiere)
    if promotion_id is not None:
        query = query.filter(Utilisateur.promotion_id == promotion_id)

    rows = []
    for item in query.order_by(Utilisateur.nom.asc(), Utilisateur.prenom.asc()).all():
        promotion = item.etudiant.promotion
        rows.append([
            item.etudiant.id,
            item.etudiant.nom,
            item.etudiant.prenom,
            item.etudiant.email,
            promotion.niveau if promotion else '',
            promotion.filiere.nom if promotion else '',
            item.matiere.code,
            item.matiere.nom,
            item.nombre_absences,
            item.nb_absences_justifiees,
            item.seuil_atteint,
            item.updated_at.isoformat(),
        ])
    return rows


def _course_attendance_export_rows(cours):
    students, _summary = _course_attendance_rows(cours)
    return [
        [
            student['id'],
            student['nom'],
            student['prenom'],
            student['email'],
            student['statut_presence'],
            student['pointage']['timestamp_pointage'] if student['pointage'] else '',
            student['pointage']['methode'] if student['pointage'] else '',
            student['pointage']['justificatif'] if student['pointage'] else '',
        ]
        for student in students
    ]


def _delete_entity(instance, conflict_message):
    db.session.delete(instance)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': conflict_message,
            'details': {},
        }), 409

    return jsonify({'message': 'Suppression effectuee avec succes.'}), 200


def _request_payload():
    if request.is_json:
        return request.get_json(silent=True) or {}
    return request.form


def _payload_value(payload, key, default=None):
    value = payload.get(key, default)
    if isinstance(value, str):
        return value.strip()
    return value


def _generate_temporary_password(length=12):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def _save_user_photo(role, required=True):
    photo = request.files.get('photo')
    if photo is None or not photo.filename:
        if required:
            raise ValueError('La photo est requise.')
        return None

    if photo is None or not photo.filename:
        raise ValueError('La photo est requise.')

    extension = Path(secure_filename(photo.filename)).suffix.lower()
    if extension not in {'.png', '.jpg', '.jpeg', '.webp'}:
        raise ValueError('Le format de photo est invalide. Formats acceptes: png, jpg, jpeg, webp.')

    relative_dir = Path('users') / role
    absolute_dir = Path(current_app.config['UPLOAD_FOLDER']) / relative_dir
    absolute_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{role}_{secrets.token_hex(8)}{extension}"
    photo.save(str(absolute_dir / filename))
    return str((relative_dir / filename).as_posix())


def _delete_photo_if_exists(photo_path):
    if not photo_path:
        return
    full_path = Path(current_app.config['UPLOAD_FOLDER']) / photo_path
    if full_path.exists():
        full_path.unlink()


def _get_managed_user_or_error(user_id, expected_role):
    user = get_or_404(Utilisateur, user_id, 'Utilisateur introuvable.')
    if user.role != expected_role:
        return None, jsonify({
            'error': 'bad_request',
            'message': f"L'utilisateur demande n'est pas un {expected_role}.",
            'details': {},
        }), 400
    return user, None, None


def _student_attendance_query(etudiant_id):
    return Pointage.query.filter_by(etudiant_id=etudiant_id).order_by(
        Pointage.timestamp_pointage.desc()
    )


def _student_attendance_history_payload(etudiant, page, per_page):
    payload = paginate_query(
        _student_attendance_query(etudiant.id),
        serialize_pointage,
        page,
        per_page,
    )
    return {
        'student': serialize_utilisateur(etudiant),
        'history': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }


def _student_attendance_summary_payload(etudiant):
    pointages = _student_attendance_query(etudiant.id).all()
    summary = {
        'present': 0,
        'retard': 0,
        'absent': 0,
        'absence_justifiee': 0,
        'invalide': 0,
    }
    by_subject = {}

    for pointage in pointages:
        summary[pointage.statut] = summary.get(pointage.statut, 0) + 1
        matiere = pointage.cours.matiere
        subject_row = by_subject.setdefault(matiere.id, {
            'subject_id': matiere.id,
            'subject_code': matiere.code,
            'subject_name': matiere.nom,
            'present': 0,
            'retard': 0,
            'absent': 0,
            'absence_justifiee': 0,
            'invalide': 0,
            'total': 0,
        })
        subject_row[pointage.statut] = subject_row.get(pointage.statut, 0) + 1
        subject_row['total'] += 1

    suivis = SuiviAbsences.query.filter_by(etudiant_id=etudiant.id).all()
    return {
        'student': serialize_utilisateur(etudiant),
        'totals': {
            **summary,
            'total_pointages': len(pointages),
            'total_suivis_absences': len(suivis),
            'total_absences_injustifiees': sum(item.nombre_absences for item in suivis),
            'total_absences_justifiees': sum(item.nb_absences_justifiees for item in suivis),
            'seuils_atteints': sum(1 for item in suivis if item.seuil_atteint),
        },
        'by_subject': list(by_subject.values()),
    }


def _student_attendance_by_course_payload(etudiant):
    rows = []
    for pointage in _student_attendance_query(etudiant.id).all():
        rows.append({
            'attendance': serialize_pointage(pointage),
            'course': serialize_cours(pointage.cours, etudiant_id=etudiant.id),
        })
    return {
        'student': serialize_utilisateur(etudiant),
        'rows': rows,
        'count': len(rows),
    }


def _dashboard_summary_payload():
    total_etudiants = Utilisateur.query.filter_by(role='etudiant', statut='actif').count()
    total_profs = Utilisateur.query.filter_by(role='professeur', statut='actif').count()
    total_promotions = Promotion.query.count()
    total_filieres = Filiere.query.count()
    cours_aujourdhui = Cours.query.filter_by(date=date.today()).count()
    absences_critiques = SuiviAbsences.query.filter_by(seuil_atteint=True).count()

    repartition_promotions = []
    for promotion in Promotion.query.order_by(Promotion.niveau.asc()).all():
        repartition_promotions.append({
            'promotion_id': promotion.id,
            'niveau': promotion.niveau,
            'annee_academique': promotion.annee_academique,
            'filiere': promotion.filiere.nom,
            'students_count': len(promotion.etudiants),
            'courses_count': len(promotion.cours),
        })

    repartition_filieres = []
    for filiere in Filiere.query.order_by(Filiere.nom.asc()).all():
        promotions = filiere.promotions
        repartition_filieres.append({
            'department_id': filiere.id,
            'department_name': filiere.nom,
            'promotions_count': len(promotions),
            'students_count': sum(len(promo.etudiants) for promo in promotions),
            'courses_count': sum(len(promo.cours) for promo in promotions),
        })

    repartition_matieres = []
    for matiere in Matiere.query.order_by(Matiere.code.asc()).all():
        repartition_matieres.append({
            'subject_id': matiere.id,
            'subject_code': matiere.code,
            'subject_name': matiere.nom,
            'courses_count': len(matiere.cours),
        })

    return {
        'today': date.today().isoformat(),
        'totals': {
            'students': total_etudiants,
            'professors': total_profs,
            'promotions': total_promotions,
            'departments': total_filieres,
            'courses_today': cours_aujourdhui,
            'critical_absence_cases': absences_critiques,
        },
        'distribution': {
            'by_promotion': repartition_promotions,
            'by_department': repartition_filieres,
            'by_subject': repartition_matieres,
        },
    }


def _dashboard_trends_payload(days=7):
    daily_rows = []
    for delta in range(days - 1, -1, -1):
        current_day = date.today() - timedelta(days=delta)
        courses = Cours.query.filter_by(date=current_day).all()
        pointages = Pointage.query.join(Cours).filter(Cours.date == current_day).all()
        daily_rows.append({
            'date': current_day.isoformat(),
            'courses_count': len(courses),
            'present_count': sum(1 for item in pointages if item.statut == 'present'),
            'retard_count': sum(1 for item in pointages if item.statut == 'retard'),
            'absence_justifiee_count': sum(1 for item in pointages if item.statut == 'absence_justifiee'),
            'invalide_count': sum(1 for item in pointages if item.statut == 'invalide'),
        })

    return {
        'range_days': days,
        'generated_at': utcnow().isoformat(),
        'daily': daily_rows,
    }


def _create_managed_user(role):
    payload = _request_payload()
    nom = _payload_value(payload, 'nom')
    prenom = _payload_value(payload, 'prenom')
    email = _payload_value(payload, 'email')
    promotion_id = _payload_value(payload, 'promotion_id')

    if not nom or not prenom or not email:
        return None, ({
            'error': 'bad_request',
            'message': 'Les champs nom, prenom et email sont requis.',
            'details': {},
        }, 400)

    promotion = None
    if role == 'etudiant':
        if not promotion_id:
            return None, ({
                'error': 'bad_request',
                'message': 'Le champ promotion_id est requis pour un etudiant.',
                'details': {},
            }, 400)
        try:
            promotion = db.session.get(Promotion, int(promotion_id))
        except (TypeError, ValueError):
            promotion = None
        if promotion is None:
            return None, ({
                'error': 'bad_request',
                'message': 'La promotion fournie est invalide.',
                'details': {},
            }, 400)

    try:
        photo_path = _save_user_photo(role)
    except ValueError as exc:
        return None, ({
            'error': 'bad_request',
            'message': str(exc),
            'details': {},
        }, 400)

    temporary_password = _generate_temporary_password()
    user = Utilisateur(
        nom=nom,
        prenom=prenom,
        email=email,
        role=role,
        promotion_id=promotion.id if promotion else None,
        photo=photo_path,
        must_change_password=True,
    )
    user.set_password(temporary_password)
    db.session.add(user)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        photo_full_path = Path(current_app.config['UPLOAD_FOLDER']) / photo_path
        if photo_full_path.exists():
            photo_full_path.unlink()
        return None, ({
            'error': 'conflict',
            'message': 'Un utilisateur avec cet email existe deja.',
            'details': {},
        }, 409)

    email_sent = send_temporary_password_email(user, temporary_password)
    return {
        'user': user,
        'temporary_password': temporary_password,
        'email_sent': email_sent,
    }, None


def _update_managed_user(user):
    payload = _request_payload()
    new_photo_path = None
    previous_photo_path = user.photo

    if 'nom' in payload:
        user.nom = _payload_value(payload, 'nom')
    if 'prenom' in payload:
        user.prenom = _payload_value(payload, 'prenom')
    if 'email' in payload:
        user.email = _payload_value(payload, 'email')
    if 'statut' in payload:
        statut = _payload_value(payload, 'statut')
        if statut not in {'actif', 'inactif'}:
            return ({
                'error': 'bad_request',
                'message': "Le statut doit etre 'actif' ou 'inactif'.",
                'details': {},
            }, 400)
        user.statut = statut

    if user.role == 'etudiant' and 'promotion_id' in payload:
        try:
            promotion = db.session.get(Promotion, int(_payload_value(payload, 'promotion_id')))
        except (TypeError, ValueError):
            promotion = None
        if promotion is None:
            return ({
                'error': 'bad_request',
                'message': 'La promotion fournie est invalide.',
                'details': {},
            }, 400)
        user.promotion_id = promotion.id

    if user.role == 'etudiant' and user.promotion_id is None:
        return ({
            'error': 'bad_request',
            'message': 'Un etudiant doit etre rattache a une promotion.',
            'details': {},
        }, 400)

    if not user.nom or not user.prenom or not user.email:
        return ({
            'error': 'bad_request',
            'message': 'Les champs nom, prenom et email ne peuvent pas etre vides.',
            'details': {},
        }, 400)

    try:
        new_photo_path = _save_user_photo(user.role, required=False)
    except ValueError as exc:
        return ({
            'error': 'bad_request',
            'message': str(exc),
            'details': {},
        }, 400)

    if new_photo_path:
        user.photo = new_photo_path

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        if new_photo_path:
            _delete_photo_if_exists(new_photo_path)
            user.photo = previous_photo_path
        return ({
            'error': 'conflict',
            'message': 'Un utilisateur avec cet email existe deja.',
            'details': {},
        }, 409)

    if new_photo_path and previous_photo_path and previous_photo_path != new_photo_path:
        _delete_photo_if_exists(previous_photo_path)

    return None


def _reset_user_temporary_password(user):
    temporary_password = _generate_temporary_password()
    user.set_password(temporary_password)
    user.must_change_password = True
    db.session.commit()
    email_sent = send_temporary_password_email(user, temporary_password)
    return email_sent


def _import_users_from_excel(role):
    upload = request.files.get('file')
    if upload is None or not upload.filename:
        return None, ({
            'error': 'bad_request',
            'message': 'Le fichier Excel est requis.',
            'details': {},
        }, 400)

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        return None, ({
            'error': 'bad_request',
            'message': 'Le fichier fourni est invalide ou illisible.',
            'details': {},
        }, 400)

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None, ({
            'error': 'bad_request',
            'message': 'Le fichier Excel est vide.',
            'details': {},
        }, 400)

    headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
    required_headers = ['nom', 'prenom', 'email']
    if role == 'etudiant':
        required_headers.append('promotion_id')

    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        return None, ({
            'error': 'bad_request',
            'message': 'Le fichier Excel ne contient pas toutes les colonnes requises.',
            'details': {'missing_headers': missing_headers},
        }, 400)

    created = []
    errors = []
    email_sent_count = 0

    for row_index, row in enumerate(rows[1:], start=2):
        values = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        nom = str(values.get('nom') or '').strip()
        prenom = str(values.get('prenom') or '').strip()
        email = str(values.get('email') or '').strip()

        if not nom or not prenom or not email:
            errors.append({'row': row_index, 'message': 'Les champs nom, prenom et email sont requis.'})
            continue

        if Utilisateur.query.filter_by(email=email).first() is not None:
            errors.append({'row': row_index, 'message': 'Un utilisateur avec cet email existe deja.'})
            continue

        promotion = None
        if role == 'etudiant':
            try:
                promotion = db.session.get(Promotion, int(values.get('promotion_id')))
            except (TypeError, ValueError):
                promotion = None
            if promotion is None:
                errors.append({'row': row_index, 'message': 'La promotion fournie est invalide.'})
                continue

        temporary_password = _generate_temporary_password()
        user = Utilisateur(
            nom=nom,
            prenom=prenom,
            email=email,
            role=role,
            promotion_id=promotion.id if promotion else None,
            must_change_password=True,
        )
        user.set_password(temporary_password)
        db.session.add(user)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            errors.append({'row': row_index, 'message': 'Impossible de creer cet utilisateur.'})
            continue

        created.append(serialize_utilisateur(user))
        if send_temporary_password_email(user, temporary_password):
            email_sent_count += 1

    return {
        'created': created,
        'created_count': len(created),
        'email_sent_count': email_sent_count,
        'errors': errors,
        'errors_count': len(errors),
    }, None


@admin_bp.route('/dashboard')
@role_required('admin', 'responsable')
def dashboard():
    total_etudiants = Utilisateur.query.filter_by(role='etudiant', statut='actif').count()
    total_profs = Utilisateur.query.filter_by(role='professeur', statut='actif').count()
    cours_aujourdhui = Cours.query.filter_by(date=date.today()).count()
    total_promotions = Promotion.query.count()
    cours_recents = Cours.query.order_by(Cours.created_at.desc()).limit(5).all()

    return jsonify({
        'total_etudiants': total_etudiants,
        'total_profs': total_profs,
        'cours_aujourdhui': cours_aujourdhui,
        'total_promotions': total_promotions,
        'cours_recents': [
            {
                'id': c.id,
                'matiere': c.matiere.nom,
                'date': c.date.isoformat(),
                'heure_debut': c.heure_debut.isoformat(),
                'heure_fin': c.heure_fin.isoformat(),
                'statut': c.statut,
            }
            for c in cours_recents
        ],
        'today': date.today().isoformat(),
    }), 200


@admin_bp.route('/dashboard/summary')
@role_required('admin', 'responsable')
def dashboard_summary():
    return jsonify(_dashboard_summary_payload()), 200


@admin_bp.route('/dashboard/trends')
@role_required('admin', 'responsable')
def dashboard_trends():
    days = request.args.get('days', default=7, type=int) or 7
    days = max(1, min(days, 30))
    return jsonify(_dashboard_trends_payload(days=days)), 200


@admin_bp.route('/notifications', methods=['GET'])
@role_required('admin', 'responsable')
def list_notifications():
    notifications = Notification.query.filter_by(
        destinataire_id=int(get_jwt_identity())
    ).order_by(Notification.created_at.desc())

    page, per_page = get_pagination_params(request)
    payload = paginate_query(notifications, serialize_notification, page, per_page)
    return jsonify({
        'notifications': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
        'non_lues': Notification.query.filter_by(
            destinataire_id=int(get_jwt_identity()),
            lu=False,
        ).count(),
    }), 200


@admin_bp.route('/notifications/<int:notification_id>/read', methods=['PATCH'])
@role_required('admin', 'responsable')
def mark_notification_as_read(notification_id):
    notification = Notification.query.filter_by(
        id=notification_id,
        destinataire_id=int(get_jwt_identity()),
    ).first()
    if notification is None:
        return jsonify({
            'error': 'not_found',
            'message': 'Notification introuvable.',
            'details': {},
        }), 404

    notification.marquer_lu()
    db.session.commit()
    return jsonify({
        'message': 'Notification marquee comme lue.',
        'notification': serialize_notification(notification),
    }), 200


@admin_bp.route('/students', methods=['GET'])
@role_required('admin', 'responsable')
def list_students():
    query = Utilisateur.query.filter_by(role='etudiant')

    promotion_id = request.args.get('promotion_id', type=int)
    if promotion_id is not None:
        query = query.filter_by(promotion_id=promotion_id)

    statut = request.args.get('statut')
    if statut:
        query = query.filter_by(statut=statut)

    search = request.args.get('search')
    if search:
        pattern = f'%{search}%'
        query = query.filter(
            (Utilisateur.nom.ilike(pattern)) |
            (Utilisateur.prenom.ilike(pattern)) |
            (Utilisateur.email.ilike(pattern))
        )

    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        query.order_by(Utilisateur.nom.asc(), Utilisateur.prenom.asc()),
        serialize_utilisateur,
        page,
        per_page,
    )
    return jsonify({
        'students': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/students', methods=['POST'])
@role_required('admin', 'responsable')
def create_student():
    result, error = _create_managed_user('etudiant')
    if error is not None:
        body, status = error
        return jsonify(body), status

    response = {
        'message': 'Etudiant cree avec succes.',
        'student': serialize_utilisateur(result['user']),
        'email_sent': result['email_sent'],
    }
    if not result['email_sent']:
        response['warning'] = 'Compte cree mais email non envoye. Verifiez la configuration SMTP.'
    return jsonify(response), 201


@admin_bp.route('/professors', methods=['GET'])
@role_required('admin', 'responsable')
def list_professors():
    query = Utilisateur.query.filter_by(role='professeur')

    statut = request.args.get('statut')
    if statut:
        query = query.filter_by(statut=statut)

    search = request.args.get('search')
    if search:
        pattern = f'%{search}%'
        query = query.filter(
            (Utilisateur.nom.ilike(pattern)) |
            (Utilisateur.prenom.ilike(pattern)) |
            (Utilisateur.email.ilike(pattern))
        )

    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        query.order_by(Utilisateur.nom.asc(), Utilisateur.prenom.asc()),
        serialize_utilisateur,
        page,
        per_page,
    )
    return jsonify({
        'professors': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/professors', methods=['POST'])
@role_required('admin', 'responsable')
def create_professor():
    result, error = _create_managed_user('professeur')
    if error is not None:
        body, status = error
        return jsonify(body), status

    response = {
        'message': 'Professeur cree avec succes.',
        'professor': serialize_utilisateur(result['user']),
        'email_sent': result['email_sent'],
    }
    if not result['email_sent']:
        response['warning'] = 'Compte cree mais email non envoye. Verifiez la configuration SMTP.'
    return jsonify(response), 201


@admin_bp.route('/students/import', methods=['POST'])
@role_required('admin', 'responsable')
def import_students():
    result, error = _import_users_from_excel('etudiant')
    if error is not None:
        body, status = error
        return jsonify(body), status

    return jsonify({
        'message': 'Import des etudiants termine.',
        **result,
    }), 200


@admin_bp.route('/professors/import', methods=['POST'])
@role_required('admin', 'responsable')
def import_professors():
    result, error = _import_users_from_excel('professeur')
    if error is not None:
        body, status = error
        return jsonify(body), status

    return jsonify({
        'message': 'Import des professeurs termine.',
        **result,
    }), 200


@admin_bp.route('/students/<int:etudiant_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_student(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    suivis = SuiviAbsences.query.filter_by(etudiant_id=etudiant.id).all()
    pointages = Pointage.query.filter_by(etudiant_id=etudiant.id).order_by(
        Pointage.timestamp_pointage.desc()
    ).limit(20).all()

    return jsonify({
        'student': serialize_utilisateur(etudiant),
        'absence_summary': [serialize_suivi_absence(item) for item in suivis],
        'recent_attendance': [
            {
                'id': item.id,
                'cours_id': item.cours_id,
                'statut': item.statut,
                'timestamp_pointage': item.timestamp_pointage.isoformat(),
                'matiere': item.cours.matiere.nom,
                'date_cours': item.cours.date.isoformat(),
            }
            for item in pointages
        ],
    }), 200


@admin_bp.route('/students/<int:etudiant_id>/attendance/history', methods=['GET'])
@role_required('admin', 'responsable')
def get_student_attendance_history(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    page, per_page = get_pagination_params(request)
    return jsonify(_student_attendance_history_payload(etudiant, page, per_page)), 200


@admin_bp.route('/students/<int:etudiant_id>/attendance/summary', methods=['GET'])
@role_required('admin', 'responsable')
def get_student_attendance_summary(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    return jsonify(_student_attendance_summary_payload(etudiant)), 200


@admin_bp.route('/students/<int:etudiant_id>/attendance/by-course', methods=['GET'])
@role_required('admin', 'responsable')
def get_student_attendance_by_course(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    return jsonify(_student_attendance_by_course_payload(etudiant)), 200


@admin_bp.route('/students/<int:etudiant_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_student(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    error = _update_managed_user(etudiant)
    if error is not None:
        body, status = error
        return jsonify(body), status

    return jsonify({
        'message': 'Etudiant mis a jour avec succes.',
        'student': serialize_utilisateur(etudiant),
    }), 200


@admin_bp.route('/students/<int:etudiant_id>/reset-password', methods=['POST'])
@role_required('admin', 'responsable')
def reset_student_password(etudiant_id):
    etudiant, error_response, status_code = _get_managed_user_or_error(etudiant_id, 'etudiant')
    if error_response is not None:
        return error_response, status_code

    email_sent = _reset_user_temporary_password(etudiant)
    response = {
        'message': 'Mot de passe temporaire regenere avec succes.',
        'student': serialize_utilisateur(etudiant),
        'email_sent': email_sent,
    }
    if not email_sent:
        response['warning'] = 'Mot de passe regenere mais email non envoye.'
    return jsonify(response), 200


@admin_bp.route('/professors/<int:professeur_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_professor(professeur_id):
    professeur, error_response, status_code = _get_managed_user_or_error(professeur_id, 'professeur')
    if error_response is not None:
        return error_response, status_code

    cours = Cours.query.filter_by(professeur_id=professeur.id).order_by(
        Cours.date.desc(),
        Cours.heure_debut.desc(),
    ).limit(20).all()

    return jsonify({
        'professor': serialize_utilisateur(professeur),
        'recent_courses': [serialize_cours(item) for item in cours],
    }), 200


@admin_bp.route('/professors/<int:professeur_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_professor(professeur_id):
    professeur, error_response, status_code = _get_managed_user_or_error(professeur_id, 'professeur')
    if error_response is not None:
        return error_response, status_code

    error = _update_managed_user(professeur)
    if error is not None:
        body, status = error
        return jsonify(body), status

    return jsonify({
        'message': 'Professeur mis a jour avec succes.',
        'professor': serialize_utilisateur(professeur),
    }), 200


@admin_bp.route('/professors/<int:professeur_id>/reset-password', methods=['POST'])
@role_required('admin', 'responsable')
def reset_professor_password(professeur_id):
    professeur, error_response, status_code = _get_managed_user_or_error(professeur_id, 'professeur')
    if error_response is not None:
        return error_response, status_code

    email_sent = _reset_user_temporary_password(professeur)
    response = {
        'message': 'Mot de passe temporaire regenere avec succes.',
        'professor': serialize_utilisateur(professeur),
        'email_sent': email_sent,
    }
    if not email_sent:
        response['warning'] = 'Mot de passe regenere mais email non envoye.'
    return jsonify(response), 200


@admin_bp.route('/promotions', methods=['GET'])
@role_required('admin', 'responsable')
def list_promotions():
    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        Promotion.query.order_by(Promotion.annee_academique.desc(), Promotion.niveau.asc()),
        lambda promo: {
            **serialize_promotion(promo),
            'nombre_etudiants': len(promo.etudiants),
            'nombre_cours': len(promo.cours),
        },
        page,
        per_page,
    )
    return jsonify({
        'promotions': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/courses', methods=['GET'])
@role_required('admin', 'responsable')
def list_courses():
    query = Cours.query

    promotion_id = request.args.get('promotion_id', type=int)
    if promotion_id is not None:
        query = query.filter_by(promotion_id=promotion_id)

    professeur_id = request.args.get('professeur_id', type=int)
    if professeur_id is not None:
        query = query.filter_by(professeur_id=professeur_id)

    statut = request.args.get('statut')
    if statut:
        query = query.filter_by(statut=statut)

    date_cours = request.args.get('date')
    if date_cours:
        try:
            query = query.filter_by(date=datetime.strptime(date_cours, '%Y-%m-%d').date())
        except ValueError:
            return jsonify({
                'error': 'bad_request',
                'message': 'Le paramètre date doit être au format YYYY-MM-DD.',
                'details': {},
            }), 400

    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        query.order_by(Cours.date.asc(), Cours.heure_debut.asc()),
        serialize_cours,
        page,
        per_page,
    )
    return jsonify({
        'courses': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/courses', methods=['POST'])
@role_required('admin', 'responsable')
def create_course():
    data = request.get_json(silent=True) or {}
    aliases = {
        'matiere_id': data.get('matiere_id') or data.get('subject_id'),
        'professeur_id': data.get('professeur_id') or data.get('professor_id'),
        'salle_id': data.get('salle_id') or data.get('room_id'),
        'promotion_id': data.get('promotion_id'),
        'date': data.get('date'),
        'heure_debut': data.get('heure_debut') or data.get('start_time'),
        'heure_fin': data.get('heure_fin') or data.get('end_time'),
        'tolerance_retard_minutes': data.get('tolerance_retard_minutes'),
        'statut': data.get('statut', 'programme'),
        'motif_changement': data.get('motif_changement') or data.get('reason'),
    }
    required_fields = [
        'matiere_id', 'professeur_id', 'salle_id', 'promotion_id',
        'date', 'heure_debut', 'heure_fin',
    ]
    missing = [field for field in required_fields if not aliases.get(field)]
    if missing:
        return jsonify({
            'error': 'bad_request',
            'message': 'Certains champs obligatoires sont manquants.',
            'details': {
                'missing_fields': missing,
                'accepted_fields': {
                    'matiere_id': ['matiere_id', 'subject_id'],
                    'professeur_id': ['professeur_id', 'professor_id'],
                    'salle_id': ['salle_id', 'room_id'],
                    'heure_debut': ['heure_debut', 'start_time'],
                    'heure_fin': ['heure_fin', 'end_time'],
                    'motif_changement': ['motif_changement', 'reason'],
                },
            },
        }), 400

    try:
        date_cours = datetime.strptime(aliases['date'], '%Y-%m-%d').date()
        heure_debut = datetime.strptime(aliases['heure_debut'], '%H:%M:%S').time()
        heure_fin = datetime.strptime(aliases['heure_fin'], '%H:%M:%S').time()
    except ValueError:
        return jsonify({
            'error': 'bad_request',
            'message': 'Les formats attendus sont date=YYYY-MM-DD et heure=HH:MM:SS.',
            'details': {},
        }), 400

    if heure_fin <= heure_debut:
        return jsonify({
            'error': 'bad_request',
            'message': "L'heure de fin doit être postérieure à l'heure de début.",
            'details': {},
        }), 400

    matiere = db.session.get(Matiere, aliases['matiere_id'])
    professeur = db.session.get(Utilisateur, aliases['professeur_id'])
    salle = db.session.get(Salle, aliases['salle_id'])
    promotion = db.session.get(Promotion, aliases['promotion_id'])

    if None in (matiere, professeur, salle, promotion):
        return jsonify({
            'error': 'bad_request',
            'message': 'Une ou plusieurs références fournies sont invalides.',
            'details': {
                'matiere_id': bool(matiere),
                'professeur_id': bool(professeur),
                'salle_id': bool(salle),
                'promotion_id': bool(promotion),
            },
        }), 400

    if professeur.role != 'professeur':
        return jsonify({
            'error': 'bad_request',
            'message': "L'utilisateur sélectionné doit avoir le rôle professeur.",
            'details': {},
        }), 400

    tolerance_retard_minutes = aliases['tolerance_retard_minutes']
    if tolerance_retard_minutes is None:
        tolerance_retard_minutes = Parametre.get_solo().tolerance_retard_minutes_defaut

    cours = Cours(
        matiere_id=matiere.id,
        professeur_id=professeur.id,
        salle_id=salle.id,
        promotion_id=promotion.id,
        date=date_cours,
        heure_debut=heure_debut,
        heure_fin=heure_fin,
        tolerance_retard_minutes=tolerance_retard_minutes,
        statut=aliases['statut'],
        motif_changement=aliases['motif_changement'],
        created_by=int(get_jwt_identity()),
    )
    db.session.add(cours)
    db.session.commit()

    return jsonify({
        'message': 'Cours créé avec succès.',
        'course': serialize_cours(cours),
    }), 201


@admin_bp.route('/courses/<int:cours_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_course_details(cours_id):
    cours = get_or_404(Cours, cours_id, 'Cours introuvable.')
    students, attendance_summary = _course_attendance_rows(cours)

    return jsonify({
        'course': serialize_cours(cours),
        'attendance_summary': attendance_summary,
        'students': students,
        'notifications_count': Notification.query.filter_by(cours_id=cours.id).count(),
    }), 200


@admin_bp.route('/absences', methods=['GET'])
@role_required('admin', 'responsable')
def list_absence_tracking():
    query = SuiviAbsences.query

    etudiant_id = request.args.get('etudiant_id', type=int)
    if etudiant_id is not None:
        query = query.filter_by(etudiant_id=etudiant_id)

    seuil_atteint = request.args.get('seuil_atteint')
    if seuil_atteint in ('true', 'false'):
        query = query.filter_by(seuil_atteint=(seuil_atteint == 'true'))

    promotion_id = request.args.get('promotion_id', type=int)
    if promotion_id is not None:
        query = query.join(SuiviAbsences.etudiant).filter(Utilisateur.promotion_id == promotion_id)

    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        query.order_by(
            SuiviAbsences.seuil_atteint.desc(),
            SuiviAbsences.nombre_absences.desc(),
            SuiviAbsences.updated_at.desc(),
        ),
        lambda item: {
            **serialize_suivi_absence(item),
            'etudiant': {
                'id': item.etudiant.id,
                'nom': item.etudiant.nom,
                'prenom': item.etudiant.prenom,
                'email': item.etudiant.email,
            },
            'promotion': {
                'id': item.etudiant.promotion.id,
                'niveau': item.etudiant.promotion.niveau,
                'annee_academique': item.etudiant.promotion.annee_academique,
                'filiere': item.etudiant.promotion.filiere.nom,
            } if item.etudiant.promotion else None,
        },
        page,
        per_page,
    )
    return jsonify({
        'absence_tracking': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/absences/<int:suivi_id>/justify', methods=['POST'])
@role_required('admin', 'responsable')
def justify_absence(suivi_id):
    suivi = get_or_404(SuiviAbsences, suivi_id, 'Suivi d absences introuvable.')

    if suivi.nombre_absences <= 0:
        return jsonify({
            'error': 'bad_request',
            'message': "Aucune absence injustifiee n'est disponible a justifier.",
            'details': {},
        }), 400

    payload = request.get_json(silent=True) or {}
    justification = {
        'reason': (payload.get('reason') or payload.get('justificatif') or '').strip() or None,
        'document_url': (payload.get('document_url') or '').strip() or None,
        'status': (payload.get('status') or 'justified').strip() or 'justified',
    }

    suivi.justifier()
    db.session.commit()

    return jsonify({
        'message': 'Absence justifiee avec succes.',
        'absence_tracking': serialize_suivi_absence(suivi),
        'justification': justification,
    }), 200


@admin_bp.route('/matieres', methods=['GET'])
@admin_bp.route('/subjects', methods=['GET'])
@role_required('admin', 'responsable')
def list_matieres():
    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        Matiere.query.order_by(Matiere.code.asc()),
        serialize_matiere,
        page,
        per_page,
    )
    return jsonify({
        'subjects': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/matieres', methods=['POST'])
@admin_bp.route('/subjects', methods=['POST'])
@role_required('admin', 'responsable')
def create_matiere():
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    code = (data.get('code') or '').strip()
    credits = data.get('credits', 0)

    if not nom or not code:
        return jsonify({
            'error': 'bad_request',
            'message': 'Les champs nom et code sont requis.',
            'details': {},
        }), 400

    matiere = Matiere(nom=nom, code=code, credits=credits)
    db.session.add(matiere)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': 'Une matiere avec ce code existe deja.',
            'details': {},
        }), 409

    return jsonify({'message': 'Matiere creee avec succes.', 'subject': serialize_matiere(matiere)}), 201


@admin_bp.route('/matieres/<int:matiere_id>', methods=['GET'])
@admin_bp.route('/subjects/<int:matiere_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_matiere(matiere_id):
    matiere = get_or_404(Matiere, matiere_id, 'Matiere introuvable.')
    return jsonify({'subject': serialize_matiere(matiere)}), 200


@admin_bp.route('/matieres/<int:matiere_id>', methods=['PATCH'])
@admin_bp.route('/subjects/<int:matiere_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_matiere(matiere_id):
    matiere = get_or_404(Matiere, matiere_id, 'Matiere introuvable.')
    data = request.get_json(silent=True) or {}

    if 'nom' in data:
        matiere.nom = (data.get('nom') or '').strip()
    if 'code' in data:
        matiere.code = (data.get('code') or '').strip()
    if 'credits' in data:
        matiere.credits = data.get('credits', 0)

    if not matiere.nom or not matiere.code:
        return jsonify({
            'error': 'bad_request',
            'message': 'Les champs nom et code ne peuvent pas etre vides.',
            'details': {},
        }), 400

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': 'Une matiere avec ce code existe deja.',
            'details': {},
        }), 409

    return jsonify({'message': 'Matiere mise a jour avec succes.', 'subject': serialize_matiere(matiere)}), 200


@admin_bp.route('/matieres/<int:matiere_id>', methods=['DELETE'])
@admin_bp.route('/subjects/<int:matiere_id>', methods=['DELETE'])
@role_required('admin', 'responsable')
def delete_matiere(matiere_id):
    matiere = get_or_404(Matiere, matiere_id, 'Matiere introuvable.')
    return _delete_entity(matiere, 'Impossible de supprimer cette matiere car elle est deja utilisee.')


@admin_bp.route('/salles', methods=['GET'])
@admin_bp.route('/rooms', methods=['GET'])
@role_required('admin', 'responsable')
def list_salles():
    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        Salle.query.order_by(Salle.nom.asc()),
        serialize_salle,
        page,
        per_page,
    )
    return jsonify({
        'rooms': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/salles', methods=['POST'])
@admin_bp.route('/rooms', methods=['POST'])
@role_required('admin', 'responsable')
def create_salle():
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    batiment = data.get('batiment')

    if not nom:
        return jsonify({
            'error': 'bad_request',
            'message': 'Le champ nom est requis.',
            'details': {},
        }), 400

    salle = Salle(nom=nom, batiment=batiment)
    db.session.add(salle)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': 'Une salle avec ce QR code existe deja.',
            'details': {},
        }), 409

    return jsonify({'message': 'Salle creee avec succes.', 'room': serialize_salle(salle)}), 201


@admin_bp.route('/salles/<int:salle_id>', methods=['GET'])
@admin_bp.route('/rooms/<int:salle_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_salle(salle_id):
    salle = get_or_404(Salle, salle_id, 'Salle introuvable.')
    return jsonify({'room': serialize_salle(salle)}), 200


@admin_bp.route('/salles/<int:salle_id>', methods=['PATCH'])
@admin_bp.route('/rooms/<int:salle_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_salle(salle_id):
    salle = get_or_404(Salle, salle_id, 'Salle introuvable.')
    data = request.get_json(silent=True) or {}

    if 'nom' in data:
        salle.nom = (data.get('nom') or '').strip()
    if 'batiment' in data:
        salle.batiment = data.get('batiment')

    if not salle.nom:
        return jsonify({
            'error': 'bad_request',
            'message': 'Le champ nom ne peut pas etre vide.',
            'details': {},
        }), 400

    db.session.commit()
    return jsonify({'message': 'Salle mise a jour avec succes.', 'room': serialize_salle(salle)}), 200


@admin_bp.route('/salles/<int:salle_id>', methods=['DELETE'])
@admin_bp.route('/rooms/<int:salle_id>', methods=['DELETE'])
@role_required('admin', 'responsable')
def delete_salle(salle_id):
    salle = get_or_404(Salle, salle_id, 'Salle introuvable.')
    return _delete_entity(salle, 'Impossible de supprimer cette salle car elle est deja utilisee.')


@admin_bp.route('/filieres', methods=['GET'])
@role_required('admin', 'responsable')
def list_filieres():
    page, per_page = get_pagination_params(request)
    payload = paginate_query(
        Filiere.query.order_by(Filiere.nom.asc()),
        serialize_filiere,
        page,
        per_page,
    )
    return jsonify({
        'departments': payload['items'],
        'count': payload['count'],
        'total': payload['total'],
        'pagination': payload['pagination'],
    }), 200


@admin_bp.route('/filieres', methods=['POST'])
@role_required('admin', 'responsable')
def create_filiere():
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    if not nom:
        return jsonify({
            'error': 'bad_request',
            'message': 'Le champ nom est requis.',
            'details': {},
        }), 400

    filiere = Filiere(nom=nom)
    db.session.add(filiere)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': 'Une filiere avec ce nom existe deja.',
            'details': {},
        }), 409

    return jsonify({'message': 'Filiere creee avec succes.', 'department': serialize_filiere(filiere)}), 201


@admin_bp.route('/filieres/<int:filiere_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_filiere(filiere_id):
    filiere = get_or_404(Filiere, filiere_id, 'Filiere introuvable.')
    return jsonify({'department': serialize_filiere(filiere)}), 200


@admin_bp.route('/filieres/<int:filiere_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_filiere(filiere_id):
    filiere = get_or_404(Filiere, filiere_id, 'Filiere introuvable.')
    data = request.get_json(silent=True) or {}
    if 'nom' in data:
        filiere.nom = (data.get('nom') or '').strip()

    if not filiere.nom:
        return jsonify({
            'error': 'bad_request',
            'message': 'Le champ nom ne peut pas etre vide.',
            'details': {},
        }), 400

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'error': 'conflict',
            'message': 'Une filiere avec ce nom existe deja.',
            'details': {},
        }), 409

    return jsonify({'message': 'Filiere mise a jour avec succes.', 'department': serialize_filiere(filiere)}), 200


@admin_bp.route('/filieres/<int:filiere_id>', methods=['DELETE'])
@role_required('admin', 'responsable')
def delete_filiere(filiere_id):
    filiere = get_or_404(Filiere, filiere_id, 'Filiere introuvable.')
    return _delete_entity(filiere, 'Impossible de supprimer cette filiere car elle est deja utilisee.')


@admin_bp.route('/promotions', methods=['POST'])
@role_required('admin', 'responsable')
def create_promotion():
    data = request.get_json(silent=True) or {}
    niveau = (data.get('niveau') or '').strip()
    annee_academique = (data.get('annee_academique') or '').strip()
    filiere_id = data.get('filiere_id')

    if not niveau or not annee_academique or not filiere_id:
        return jsonify({
            'error': 'bad_request',
            'message': 'Les champs niveau, annee_academique et filiere_id sont requis.',
            'details': {},
        }), 400

    filiere = db.session.get(Filiere, filiere_id)
    if filiere is None:
        return jsonify({
            'error': 'bad_request',
            'message': 'La filiere fournie est invalide.',
            'details': {},
        }), 400

    promotion = Promotion(
        niveau=niveau,
        annee_academique=annee_academique,
        filiere_id=filiere.id,
    )
    db.session.add(promotion)
    db.session.commit()

    return jsonify({'message': 'Promotion creee avec succes.', 'promotion': serialize_promotion(promotion)}), 201


@admin_bp.route('/promotions/<int:promotion_id>', methods=['GET'])
@role_required('admin', 'responsable')
def get_promotion(promotion_id):
    promotion = get_or_404(Promotion, promotion_id, 'Promotion introuvable.')
    return jsonify({
        'promotion': {
            **serialize_promotion(promotion),
            'nombre_etudiants': len(promotion.etudiants),
            'nombre_cours': len(promotion.cours),
        }
    }), 200


@admin_bp.route('/promotions/<int:promotion_id>', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_promotion(promotion_id):
    promotion = get_or_404(Promotion, promotion_id, 'Promotion introuvable.')
    data = request.get_json(silent=True) or {}

    if 'niveau' in data:
        promotion.niveau = (data.get('niveau') or '').strip()
    if 'annee_academique' in data:
        promotion.annee_academique = (data.get('annee_academique') or '').strip()
    if 'filiere_id' in data:
        filiere = db.session.get(Filiere, data.get('filiere_id'))
        if filiere is None:
            return jsonify({
                'error': 'bad_request',
                'message': 'La filiere fournie est invalide.',
                'details': {},
            }), 400
        promotion.filiere_id = filiere.id

    if not promotion.niveau or not promotion.annee_academique:
        return jsonify({
            'error': 'bad_request',
            'message': 'Les champs niveau et annee_academique ne peuvent pas etre vides.',
            'details': {},
        }), 400

    db.session.commit()
    return jsonify({'message': 'Promotion mise a jour avec succes.', 'promotion': serialize_promotion(promotion)}), 200


@admin_bp.route('/promotions/<int:promotion_id>', methods=['DELETE'])
@role_required('admin', 'responsable')
def delete_promotion(promotion_id):
    promotion = get_or_404(Promotion, promotion_id, 'Promotion introuvable.')
    return _delete_entity(promotion, 'Impossible de supprimer cette promotion car elle est deja utilisee.')


@admin_bp.route('/exports/absences', methods=['GET'])
@role_required('admin', 'responsable')
def export_absences_csv():
    promotion_id = request.args.get('promotion_id', type=int)
    rows = _absences_summary_export_rows(promotion_id)
    return build_csv_response('suivi_absences.csv', _ABSENCES_EXPORT_HEADERS, rows)


@admin_bp.route('/exports/courses/<int:cours_id>/attendance', methods=['GET'])
@role_required('admin', 'responsable')
def export_course_attendance_csv(cours_id):
    cours = get_or_404(Cours, cours_id, 'Cours introuvable.')
    rows = _course_attendance_export_rows(cours)
    filename = f"cours_{cours.id}_presence.csv"
    return build_csv_response(filename, _COURSE_ATTENDANCE_EXPORT_HEADERS, rows)


@admin_bp.route('/exports/absences/xlsx', methods=['GET'])
@role_required('admin', 'responsable')
def export_absences_excel():
    promotion_id = request.args.get('promotion_id', type=int)
    rows = _absences_summary_export_rows(promotion_id)
    return build_excel_response('suivi_absences.xlsx', 'Absences', _ABSENCES_EXPORT_HEADERS, rows)


@admin_bp.route('/exports/courses/<int:cours_id>/attendance/xlsx', methods=['GET'])
@role_required('admin', 'responsable')
def export_course_attendance_excel(cours_id):
    cours = get_or_404(Cours, cours_id, 'Cours introuvable.')
    rows = _course_attendance_export_rows(cours)
    filename = f'cours_{cours.id}_presence.xlsx'
    return build_excel_response(filename, 'Presence', _COURSE_ATTENDANCE_EXPORT_HEADERS, rows)


@admin_bp.route('/settings', methods=['GET'])
@role_required('admin', 'responsable')
def get_settings():
    parametre = Parametre.get_solo()
    return jsonify({'settings': serialize_parametre(parametre)}), 200


@admin_bp.route('/settings', methods=['PATCH'])
@role_required('admin', 'responsable')
def update_settings():
    parametre = Parametre.get_solo()
    data = request.get_json(silent=True) or {}
    seuil_modifie = False

    if 'nom_etablissement' in data:
        nom = (data.get('nom_etablissement') or '').strip()
        if not nom:
            return jsonify({
                'error': 'bad_request',
                'message': "Le nom de l etablissement ne peut pas etre vide.",
                'details': {},
            }), 400
        parametre.nom_etablissement = nom

    if 'seuil_absences' in data:
        try:
            seuil = int(data.get('seuil_absences'))
        except (TypeError, ValueError):
            seuil = None
        if seuil is None or seuil < 1:
            return jsonify({
                'error': 'bad_request',
                'message': 'seuil_absences doit etre un entier superieur ou egal a 1.',
                'details': {},
            }), 400
        seuil_modifie = seuil != parametre.seuil_absences
        parametre.seuil_absences = seuil

    if 'tolerance_retard_minutes_defaut' in data:
        try:
            tolerance = int(data.get('tolerance_retard_minutes_defaut'))
        except (TypeError, ValueError):
            tolerance = None
        if tolerance is None or tolerance < 0:
            return jsonify({
                'error': 'bad_request',
                'message': 'tolerance_retard_minutes_defaut doit etre un entier positif ou nul.',
                'details': {},
            }), 400
        parametre.tolerance_retard_minutes_defaut = tolerance

    if 'contact_support_email' in data:
        parametre.contact_support_email = (data.get('contact_support_email') or '').strip() or None

    db.session.commit()

    if seuil_modifie:
        # Le seuil pilote seuil_atteint sur tous les suivis existants : on les recalcule
        # immediatement pour eviter un etat incoherent entre l'ancien et le nouveau seuil.
        for suivi in SuiviAbsences.query.all():
            suivi.verifier_seuil()
        db.session.commit()

    return jsonify({
        'message': 'Parametres mis a jour avec succes.',
        'settings': serialize_parametre(parametre),
    }), 200


@admin_bp.route('/report-templates', methods=['GET'])
@role_required('admin', 'responsable')
def list_report_templates():
    return jsonify({'templates': REPORT_TEMPLATES, 'count': len(REPORT_TEMPLATES)}), 200


@admin_bp.route('/reports/generate', methods=['POST'])
@role_required('admin', 'responsable')
def generate_report():
    data = request.get_json(silent=True) or {}
    template_id = (data.get('template_id') or data.get('template') or '').strip()
    report_format = (data.get('format') or 'csv').strip().lower()

    if report_format not in ('csv', 'xlsx'):
        return jsonify({
            'error': 'bad_request',
            'message': "Le champ format doit etre 'csv' ou 'xlsx'.",
            'details': {},
        }), 400

    if template_id == 'absences_summary':
        promotion_id = data.get('promotion_id')
        rows = _absences_summary_export_rows(promotion_id)
        filename = f'suivi_absences.{report_format}'
        if report_format == 'xlsx':
            return build_excel_response(filename, 'Absences', _ABSENCES_EXPORT_HEADERS, rows)
        return build_csv_response(filename, _ABSENCES_EXPORT_HEADERS, rows)

    if template_id == 'course_attendance':
        cours_id = data.get('cours_id')
        if not cours_id:
            return jsonify({
                'error': 'bad_request',
                'message': 'cours_id est requis pour ce modele de rapport.',
                'details': {},
            }), 400
        cours = db.session.get(Cours, cours_id)
        if cours is None:
            return jsonify({
                'error': 'bad_request',
                'message': 'Le cours fourni est invalide.',
                'details': {},
            }), 400
        rows = _course_attendance_export_rows(cours)
        filename = f'cours_{cours.id}_presence.{report_format}'
        if report_format == 'xlsx':
            return build_excel_response(filename, 'Presence', _COURSE_ATTENDANCE_EXPORT_HEADERS, rows)
        return build_csv_response(filename, _COURSE_ATTENDANCE_EXPORT_HEADERS, rows)

    return jsonify({
        'error': 'bad_request',
        'message': 'template_id inconnu.',
        'details': {'available_templates': [template['id'] for template in REPORT_TEMPLATES]},
    }), 400
